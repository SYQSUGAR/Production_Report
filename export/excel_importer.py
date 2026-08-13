"""将标准 XLSX 或 Excel 2003 SpreadsheetML 转换为编辑器模板。"""

from datetime import date, datetime, time
from pathlib import Path
import xml.etree.ElementTree as ET

import openpyxl
try:
    from openpyxl.styles.colors import COLOR_INDEXED as _INDEXED_COLORS
except ImportError:  # openpyxl 3.0.x
    from openpyxl.styles.colors import COLOR_INDEX as _INDEXED_COLORS
from openpyxl.utils import column_index_from_string
from PyQt6.QtCore import Qt

from models.template_model import CellData, CellStyle, TemplateModel


_BORDER_STYLE_MAP = {
    "thin": ("solid", 1), "hair": ("solid", 1), "continuous": ("solid", 1),
    "medium": ("solid", 2), "thick": ("solid", 3),
    "dashed": ("dashed", 1), "mediumdashed": ("dashed", 2),
    "dotted": ("dotted", 1), "dashdot": ("dash_dot", 1),
    "mediumdashdot": ("dash_dot", 2), "dashdotdot": ("dash_dot", 1),
    "mediumdashdotdot": ("dash_dot", 2), "double": ("double", 1),
}
_SS = "urn:schemas-microsoft-com:office:spreadsheet"
_NS = {"ss": _SS}


def _a(element, name, default=None):
    return element.attrib.get(f"{{{_SS}}}{name}", default) if element is not None else default


def _color_to_hex(color) -> str | None:
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return f"#{color.rgb[-6:].upper()}"
    if color.type == "indexed" and color.indexed is not None:
        try:
            return f"#{_INDEXED_COLORS[color.indexed][-6:].upper()}"
        except (IndexError, TypeError):
            return None
    if color.type == "theme":
        return "#FFFFFF" if color.theme == 0 else "#000000" if color.theme == 1 else None
    return None


def _display_value(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if cell.data_type == "f":
        return str(value) if str(value).startswith("=") else f"={value}"
    return str(value)


def _alignment(value, vertical=False) -> int | None:
    value = (value or "").lower()
    if vertical:
        return {"top": int(Qt.AlignmentFlag.AlignTop), "center": int(Qt.AlignmentFlag.AlignVCenter),
                "bottom": int(Qt.AlignmentFlag.AlignBottom)}.get(value)
    return {"left": int(Qt.AlignmentFlag.AlignLeft), "center": int(Qt.AlignmentFlag.AlignCenter),
            "centercontinuous": int(Qt.AlignmentFlag.AlignCenter), "right": int(Qt.AlignmentFlag.AlignRight),
            "fill": int(Qt.AlignmentFlag.AlignLeft), "distributed": int(Qt.AlignmentFlag.AlignCenter),
            "justify": int(Qt.AlignmentFlag.AlignLeft)}.get(value)


def _number_format(fmt: str) -> str:
    """将 openpyxl 数字格式映射到编辑器内部键，未知格式直接透传原字符串。"""
    normalized = (fmt or "General").lower()
    if normalized == "general": return "general"
    if normalized == "@": return "text"
    if "%" in normalized: return "percent"
    if any(token in normalized for token in ("yy", "dd", "mm-", "m/", "d/")): return "date"
    if normalized == "#,##0.000" or normalized == "0.000": return "decimal_3"
    if any(token in normalized for token in (".0", ".#")): return "decimal_2"
    if any(token in normalized for token in ("0", "#")): return "integer"
    # 未知格式直接透传（如 "0.0000", "¥#,##0.00" 等自定义格式）
    return fmt if fmt and fmt != "General" else "general"


def _side(side):
    if not side or not side.style:
        return None, None, None
    style, width = _BORDER_STYLE_MAP.get(side.style.lower(), ("solid", 1))
    return style, width, _color_to_hex(side.color)


def _xlsx_cell_style(cell) -> CellStyle:
    sides = [_side(cell.border.top), _side(cell.border.bottom), _side(cell.border.left), _side(cell.border.right)]
    active = [side for side in sides if side[0]]
    fill_color = _color_to_hex(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None
    return CellStyle(
        font_family=cell.font.name, font_size=round(cell.font.sz) if cell.font.sz else None,
        bold=bool(cell.font.b), italic=bool(cell.font.i), underline=bool(cell.font.u and cell.font.u != "none"),
        alignment=_alignment(cell.alignment.horizontal),
        vertical_alignment=_alignment(cell.alignment.vertical, True),
        bg_color=fill_color, fg_color=_color_to_hex(cell.font.color),
        border_top=sides[0][0], border_bottom=sides[1][0], border_left=sides[2][0], border_right=sides[3][0],
        border_top_color=sides[0][2], border_bottom_color=sides[1][2],
        border_left_color=sides[2][2], border_right_color=sides[3][2],
        border_line_style=max(active, key=lambda x: x[1] or 1)[0] if active else None,
        border_width=max((x[1] or 1 for x in active), default=None),
        number_format=_number_format(cell.number_format),
    )


def _xml_style(element, parent=None) -> CellStyle:
    style = parent.clone() if parent else CellStyle()
    font, alignment = element.find("ss:Font", _NS), element.find("ss:Alignment", _NS)
    interior, number = element.find("ss:Interior", _NS), element.find("ss:NumberFormat", _NS)
    if font is not None:
        style.font_family = _a(font, "FontName") or style.font_family
        if _a(font, "Size"): style.font_size = round(float(_a(font, "Size")))
        if _a(font, "Bold") is not None: style.bold = _a(font, "Bold") == "1"
        if _a(font, "Italic") is not None: style.italic = _a(font, "Italic") == "1"
        if _a(font, "Underline") is not None: style.underline = _a(font, "Underline") != "None"
        style.fg_color = _a(font, "Color") or style.fg_color
    if alignment is not None:
        style.alignment = _alignment(_a(alignment, "Horizontal")) or style.alignment
        style.vertical_alignment = _alignment(_a(alignment, "Vertical"), True) or style.vertical_alignment
    if interior is not None and _a(interior, "Pattern") not in (None, "None"):
        style.bg_color = _a(interior, "Color") or _a(interior, "PatternColor")
    if number is not None:
        style.number_format = _number_format(_a(number, "Format", "General"))
    side_names = {"Top": "top", "Bottom": "bottom", "Left": "left", "Right": "right"}
    max_width, common_style = 0, None
    for border in element.findall("ss:Borders/ss:Border", _NS):
        side = side_names.get(_a(border, "Position"))
        if not side: continue
        line = (_a(border, "LineStyle", "Continuous") or "Continuous").lower()
        mapped, fallback = _BORDER_STYLE_MAP.get(line, ("solid", 1))
        width = max(fallback, round(float(_a(border, "Weight", fallback))))
        setattr(style, f"border_{side}", mapped)
        setattr(style, f"border_{side}_color", _a(border, "Color"))
        if width >= max_width: max_width, common_style = width, mapped
    if max_width:
        style.border_width, style.border_line_style = max_width, common_style
    return style


def _import_xlsx(filepath: str) -> TemplateModel:
    workbook = openpyxl.load_workbook(filepath, data_only=False)
    worksheet = workbook.active
    template = TemplateModel(max(worksheet.max_row, 1), max(worksheet.max_column, 1))
    template.default_style = CellStyle(
        font_family="Calibri", font_size=11,
        vertical_alignment=int(Qt.AlignmentFlag.AlignVCenter),
    )
    template.template_name = worksheet.title or Path(filepath).stem
    for row in worksheet.iter_rows():
        for cell in row:
            r, c = cell.row - 1, cell.column - 1
            value = _display_value(cell)
            if value: template.set_cell_data(r, c, CellData(static_text=value))
            if cell.has_style: template.set_cell_style(r, c, _xlsx_cell_style(cell))
    for merged in worksheet.merged_cells.ranges:
        template.add_merge_range(merged.min_row - 1, merged.max_row - 1, merged.min_col - 1, merged.max_col - 1)
    for index, dimension in worksheet.row_dimensions.items():
        if dimension.height is not None: template.row_heights[index - 1] = max(1, round(dimension.height * 96 / 72))
    for letter, dimension in worksheet.column_dimensions.items():
        if dimension.width is not None:
            col = column_index_from_string(letter) - 1
            if col < template.cols: template.col_widths[col] = max(20, min(1000, int(dimension.width * 7 + 5)))
    return template


def _import_spreadsheetml(filepath: str) -> TemplateModel:
    root = ET.parse(filepath).getroot()
    style_elements = {_a(node, "ID"): node for node in root.findall("ss:Styles/ss:Style", _NS)}
    style_cache = {}
    def resolve(style_id):
        if not style_id: return CellStyle()
        if style_id in style_cache: return style_cache[style_id]
        node = style_elements.get(style_id)
        if node is None: return CellStyle()
        parent_id = _a(node, "Parent")
        style_cache[style_id] = _xml_style(node, resolve(parent_id) if parent_id else None)
        return style_cache[style_id]

    worksheet = root.find("ss:Worksheet", _NS)
    if worksheet is None: raise ValueError("SpreadsheetML 文件中没有工作表。")
    table = worksheet.find("ss:Table", _NS)
    rows = int(_a(table, "ExpandedRowCount", "1")); cols = int(_a(table, "ExpandedColumnCount", "1"))
    template = TemplateModel(max(rows, 1), max(cols, 1))
    template.template_name = _a(worksheet, "Name", Path(filepath).stem)
    default_style = resolve("Default")
    template.default_style = default_style.clone()
    default_col_width = float(_a(table, "DefaultColumnWidth", "64"))
    default_row_height = float(_a(table, "DefaultRowHeight", "20"))
    template.col_widths = {c: max(20, round(default_col_width * 96 / 72)) for c in range(cols)}
    template.row_heights = {r: max(1, round(default_row_height * 96 / 72)) for r in range(rows)}

    current_col = 0
    for col_node in table.findall("ss:Column", _NS):
        current_col = int(_a(col_node, "Index", str(current_col + 1))) - 1
        span = int(_a(col_node, "Span", "0")); width = float(_a(col_node, "Width", str(default_col_width)))
        for col in range(current_col, min(current_col + span + 1, cols)):
            template.col_widths[col] = max(20, round(width * 96 / 72))
        current_col += span + 1

    current_row = 0
    for row_node in table.findall("ss:Row", _NS):
        current_row = int(_a(row_node, "Index", str(current_row + 1))) - 1
        if current_row >= rows: break
        if _a(row_node, "Height"): template.row_heights[current_row] = max(1, round(float(_a(row_node, "Height")) * 96 / 72))
        current_cell = 0
        for cell in row_node.findall("ss:Cell", _NS):
            current_cell = int(_a(cell, "Index", str(current_cell + 1))) - 1
            if current_cell >= cols: break
            data = cell.find("ss:Data", _NS)
            value = "" if data is None else "".join(data.itertext()).strip()
            formula = _a(cell, "Formula")
            if formula: value = formula
            if value: template.set_cell_data(current_row, current_cell, CellData(static_text=value))
            style_id = _a(cell, "StyleID") or _a(row_node, "StyleID")
            if style_id: template.set_cell_style(current_row, current_cell, resolve(style_id))
            across, down = int(_a(cell, "MergeAcross", "0")), int(_a(cell, "MergeDown", "0"))
            if across or down:
                template.add_merge_range(current_row, min(current_row + down, rows - 1),
                                         current_cell, min(current_cell + across, cols - 1))
            current_cell += across + 1
        current_row += 1
    return template


class ExcelImporter:
    @staticmethod
    def import_file(filepath: str) -> TemplateModel:
        header = Path(filepath).read_bytes()[:256].lstrip()
        if header.startswith(b"PK"):
            template = _import_xlsx(filepath)
        elif header.startswith(b"<?xml") or header.startswith(b"<Workbook"):
            template = _import_spreadsheetml(filepath)
        else:
            raise ValueError("无法识别此 Excel 文件格式；请另存为标准 .xlsx 后重试。")
        template.template_description = f"从 {Path(filepath).name} 导入"
        return template
