"""Excel 导出器——按模板内容、样式和数字格式生成最终报表。"""

from typing import Optional, Callable

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PyQt6.QtCore import Qt

from models.template_model import TemplateModel
from models.value_formatter import coerce_excel_value


_NUMBER_FORMAT_MAP = {
    "general": "General",
    "text": "@",
    "integer": "#,##0",
    "decimal_2": "#,##0.00",
    "decimal_3": "#,##0.000",
    "percent": "0.00%",
    "date": "yyyy-mm-dd",
}


class ExcelExporter:
    """根据 TemplateModel + 数据导出 Excel，完整复现模板格式。"""

    @staticmethod
    def _qt_align_to_openpyxl(alignment_int: int | None) -> str:
        if alignment_int is None:
            return "center"
        a = Qt.AlignmentFlag(alignment_int)
        if a == Qt.AlignmentFlag.AlignLeft:
            return "left"
        if a == Qt.AlignmentFlag.AlignRight:
            return "right"
        return "center"

    @staticmethod
    def _qt_valign_to_openpyxl(alignment_int: int | None) -> str:
        if alignment_int is None:
            return "center"
        a = Qt.AlignmentFlag(alignment_int)
        if a == Qt.AlignmentFlag.AlignTop:
            return "top"
        if a == Qt.AlignmentFlag.AlignBottom:
            return "bottom"
        return "center"

    @staticmethod
    def _make_border(top=None, bottom=None, left=None, right=None,
                     line_style=None, width=None) -> Border:
        style_map = {
            "solid": "thin", "dashed": "dashed", "dotted": "dotted",
            "dash_dot": "dashDot", "double": "double",
            "thin": "thin", "medium": "medium", "thick": "thick",
        }
        base_style = style_map.get(line_style or "solid", "thin")
        color = "000000"
        sides = {}
        for direction, enabled in (("left", left), ("right", right),
                                   ("top", top), ("bottom", bottom)):
            if enabled:
                sides[direction] = Side(style=base_style, color=color)
        return Border(**sides) if sides else Border()

    @classmethod
    def export(
        cls,
        template: TemplateModel,
        data: list[list[str]],
        filepath: str,
        query_callback: Optional[Callable[[int, int], Optional[str]]] = None,
    ):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = template.template_name or "Sheet1"

        if not data or template.rows == 0 or template.cols == 0:
            wb.save(filepath)
            return

        rows, cols = template.rows, template.cols

        for r_idx in range(rows):
            row_num = r_idx + 1
            for c_idx in range(cols):
                col_num = c_idx + 1
                cell_data = template.get_cell_data(r_idx, c_idx)
                cell_value = ""

                if cell_data.static_text:
                    cell_value = cell_data.static_text
                elif (cell_data.query_binding and cell_data.query_binding.enabled and query_callback):
                    db_result = query_callback(r_idx, c_idx)
                    if db_result is not None:
                        cell_value = db_result
                elif r_idx < len(data) and c_idx < len(data[r_idx]):
                    cell_value = data[r_idx][c_idx]

                # 数字格式是单元格最终输出规则，与数据来自静态文本还是数据库无关。
                style = template.get_effective_style(r_idx, c_idx)
                cell_value = coerce_excel_value(cell_value, style.number_format)
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cls._apply_style(cell, template, r_idx, c_idx)

        for mr in template.merge_ranges:
            try:
                ws.merge_cells(
                    start_row=mr.top_row + 1,
                    start_column=mr.left_col + 1,
                    end_row=mr.bottom_row + 1,
                    end_column=mr.right_col + 1,
                )
            except Exception:
                pass

        for c_idx, width in template.col_widths.items():
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = width

        for r_idx, height in template.row_heights.items():
            ws.row_dimensions[r_idx + 1].height = height

        for c_idx in range(cols):
            if c_idx in template.col_widths:
                continue
            max_len = 0
            for r_idx in range(rows):
                text = ""
                if r_idx < len(data) and c_idx < len(data[r_idx]):
                    text = str(data[r_idx][c_idx])
                cd = template.get_cell_data(r_idx, c_idx)
                if cd.static_text:
                    text = cd.static_text
                length = sum(2 if ord(ch) > 127 else 1 for ch in text)
                max_len = max(max_len, length)
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = min(max_len + 4, 50)

        wb.save(filepath)

    @classmethod
    def _apply_style(cls, cell, template: TemplateModel, row: int, col: int):
        style = template.get_effective_style(row, col)

        font_kwargs = {}
        if style.font_family:
            font_kwargs["name"] = style.font_family
        if style.font_size:
            font_kwargs["size"] = style.font_size
        if style.bold is not None:
            font_kwargs["bold"] = style.bold
        if style.italic is not None:
            font_kwargs["italic"] = style.italic
        if style.underline is not None:
            font_kwargs["underline"] = "single" if style.underline else None
        if style.fg_color:
            font_kwargs["color"] = style.fg_color.lstrip("#")
        cell.font = Font(**font_kwargs)

        cell.alignment = Alignment(
            horizontal=cls._qt_align_to_openpyxl(style.alignment),
            vertical=cls._qt_valign_to_openpyxl(style.vertical_alignment),
            wrap_text=True,
        )

        if style.bg_color:
            cell.fill = PatternFill(
                start_color=style.bg_color.lstrip("#"),
                end_color=style.bg_color.lstrip("#"),
                fill_type="solid",
            )

        cell.border = cls._make_border(
            top=style.border_top,
            bottom=style.border_bottom,
            left=style.border_left,
            right=style.border_right,
            line_style=style.border_line_style,
            width=style.border_width,
        )

        if style.number_format:
            cell.number_format = _NUMBER_FORMAT_MAP.get(style.number_format, style.number_format)
